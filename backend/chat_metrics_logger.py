"""
Enhanced Chat Metrics Logger for AI Document Intelligence System
Tracks and logs all processing metrics to Excel with historical data preservation
"""
import pandas as pd
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List
import logging
from functools import wraps
from contextlib import contextmanager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class MetricsTimer:
    """Context manager for timing operations"""
    
    def __init__(self, operation_name: str = "Operation"):
        self.operation_name = operation_name
        self.start_time = None
        self.end_time = None
        self.duration = None
    
    def __enter__(self):
        self.start_time = time.time()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.end_time = time.time()
        self.duration = self.end_time - self.start_time
        logger.debug(f"{self.operation_name} took {self.duration:.3f} seconds")
        return False
    
    def get_duration(self) -> float:
        return self.duration if self.duration else 0.0


def timing_decorator(operation_name: str = None):
    """Decorator to automatically time functions"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            name = operation_name or func.__name__
            start = time.time()
            result = func(*args, **kwargs)
            duration = time.time() - start
            logger.debug(f"{name} completed in {duration:.3f}s")
            return result, duration
        return wrapper
    return decorator


class ChatMetricsLogger:
    """
    Production-grade metrics logger for chat processing
    
    Features:
    - Single Excel file with append-only operations
    - Automatic data preservation
    - Phase-by-phase timing
    - Aggregate statistics
    - Error handling and recovery
    - Summary sheet generation
    """
    
    def __init__(self, excel_path: str = "chat_metrics_log.xlsx"):
        self.excel_path = Path(excel_path)
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Current session metrics
        self.current_metrics: Dict[str, Any] = {}
        self.session_start_time: Optional[float] = None
        
        # Column definitions
        self.columns = [
            'Timestamp',
            'Session_ID',
            'File_Name',
            'File_Type',
            'File_Size_KB',
            'Words_Count',
            'Chunks_Count',
            'Load_Time_Sec',
            'Processing_Time_Sec',
            'Embedding_Time_Sec',
            'Response_Time_Sec',
            'Total_Runtime_Sec',
            'Query_Text',
            'Response_Length',
            'Sources_Retrieved',
            'Confidence_Level',
            'Success',
            'Error_Message'
        ]
        
        # Initialize Excel file if it doesn't exist
        self._initialize_excel()
        
        logger.info(f"Metrics logger initialized: {self.excel_path}")
    
    def _initialize_excel(self):
        """Create Excel file with headers if it doesn't exist"""
        if not self.excel_path.exists():
            df = pd.DataFrame(columns=self.columns)
            
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Metrics', index=False)
                
                # Create summary sheet
                summary_df = pd.DataFrame({
                    'Metric': ['Total Sessions', 'Total Files Processed', 'Avg Processing Time', 
                              'Avg Response Time', 'Success Rate'],
                    'Value': [0, 0, 0, 0, 0]
                })
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            logger.info(f"Created new metrics file: {self.excel_path}")
    
    def start_session(self, session_id: str):
        """Start a new metrics tracking session"""
        self.session_start_time = time.time()
        self.current_metrics = {
            'session_id': session_id,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'timers': {},
            'success': True,
            'error_message': None
        }
        logger.info(f"Started metrics session: {session_id}")
    
    @contextmanager
    def track_phase(self, phase_name: str):
        """Context manager to track individual processing phases"""
        timer = MetricsTimer(phase_name)
        with timer:
            yield timer
        
        self.current_metrics['timers'][phase_name] = timer.get_duration()
    
    def log_file_info(self, file_name: str, file_type: str, file_size_bytes: int, 
                     words_count: int, chunks_count: int):
        """Log file processing information"""
        self.current_metrics.update({
            'file_name': file_name,
            'file_type': file_type,
            'file_size_kb': file_size_bytes / 1024,
            'words_count': words_count,
            'chunks_count': chunks_count
        })
    
    def log_query_info(self, query_text: str, response_length: int, 
                       sources_retrieved: int, confidence_level: str):
        """Log query and response information"""
        self.current_metrics.update({
            'query_text': query_text[:100],  # Truncate long queries
            'response_length': response_length,
            'sources_retrieved': sources_retrieved,
            'confidence_level': confidence_level
        })
    
    def log_error(self, error_message: str):
        """Log error information"""
        self.current_metrics['success'] = False
        self.current_metrics['error_message'] = error_message
    
    def end_session(self):
        """End session and write metrics to Excel"""
        if not self.session_start_time:
            logger.warning("No active session to end")
            return
        
        # Calculate total runtime
        total_runtime = time.time() - self.session_start_time
        
        # Prepare row data
        row_data = {
            'Timestamp': self.current_metrics.get('timestamp'),
            'Session_ID': self.current_metrics.get('session_id'),
            'File_Name': self.current_metrics.get('file_name', 'N/A'),
            'File_Type': self.current_metrics.get('file_type', 'N/A'),
            'File_Size_KB': self.current_metrics.get('file_size_kb', 0),
            'Words_Count': self.current_metrics.get('words_count', 0),
            'Chunks_Count': self.current_metrics.get('chunks_count', 0),
            'Load_Time_Sec': self.current_metrics['timers'].get('file_load', 0),
            'Processing_Time_Sec': self.current_metrics['timers'].get('processing', 0),
            'Embedding_Time_Sec': self.current_metrics['timers'].get('embedding', 0),
            'Response_Time_Sec': self.current_metrics['timers'].get('response_generation', 0),
            'Total_Runtime_Sec': total_runtime,
            'Query_Text': self.current_metrics.get('query_text', 'N/A'),
            'Response_Length': self.current_metrics.get('response_length', 0),
            'Sources_Retrieved': self.current_metrics.get('sources_retrieved', 0),
            'Confidence_Level': self.current_metrics.get('confidence_level', 'N/A'),
            'Success': self.current_metrics.get('success', True),
            'Error_Message': self.current_metrics.get('error_message', '')
        }
        
        # Append to Excel
        self._append_to_excel(row_data)
        
        # Update summary
        self._update_summary()
        
        # Reset session
        self.session_start_time = None
        self.current_metrics = {}
        
        logger.info(f"Session metrics saved to {self.excel_path}")
    
    def _append_to_excel(self, row_data: Dict[str, Any]):
        """Append new row to Excel file preserving existing data"""
        try:
            # Read existing data
            try:
                df = pd.read_excel(self.excel_path, sheet_name='Metrics')
            except Exception as e:
                logger.warning(f"Could not read existing data: {e}. Creating new DataFrame.")
                df = pd.DataFrame(columns=self.columns)
            
            # Append new row
            new_row = pd.DataFrame([row_data])
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Write back to Excel
            with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', 
                              if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Metrics', index=False)
                
                # Format the sheet
                self._format_metrics_sheet(writer.book['Metrics'])
            
            logger.info(f"Appended metrics row. Total rows: {len(df)}")
            
        except Exception as e:
            logger.error(f"Error appending to Excel: {e}")
            # Fallback: save to backup file
            backup_path = self.excel_path.parent / f"metrics_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            try:
                pd.DataFrame([row_data]).to_excel(backup_path, index=False)
                logger.info(f"Saved to backup file: {backup_path}")
            except:
                pass
    
    def _format_metrics_sheet(self, worksheet):
        """Apply formatting to metrics sheet"""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _update_summary(self):
        """Update summary statistics sheet"""
        try:
            # Read metrics data
            df = pd.read_excel(self.excel_path, sheet_name='Metrics')
            
            if df.empty:
                return
            
            # Calculate statistics
            total_sessions = len(df)
            total_files = df['File_Name'].nunique()
            avg_processing = df['Processing_Time_Sec'].mean()
            avg_response = df['Response_Time_Sec'].mean()
            success_rate = (df['Success'].sum() / len(df)) * 100
            
            # Additional statistics
            avg_words = df['Words_Count'].mean()
            avg_chunks = df['Chunks_Count'].mean()
            total_runtime = df['Total_Runtime_Sec'].sum()
            
            # Create summary DataFrame
            summary_data = {
                'Metric': [
                    'Total Sessions',
                    'Total Files Processed',
                    'Avg Processing Time (sec)',
                    'Avg Response Time (sec)',
                    'Success Rate (%)',
                    'Avg Words Processed',
                    'Avg Chunks Created',
                    'Total Runtime (hours)',
                    'Last Updated'
                ],
                'Value': [
                    total_sessions,
                    total_files,
                    f"{avg_processing:.3f}",
                    f"{avg_response:.3f}",
                    f"{success_rate:.1f}",
                    f"{avg_words:.0f}",
                    f"{avg_chunks:.0f}",
                    f"{total_runtime / 3600:.2f}",
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            
            # Write summary sheet
            with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', 
                              if_sheet_exists='replace') as writer:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary sheet
                ws = writer.book['Summary']
                self._format_summary_sheet(ws)
            
            logger.info("Updated summary statistics")
            
        except Exception as e:
            logger.error(f"Error updating summary: {e}")
    
    def _format_summary_sheet(self, worksheet):
        """Format the summary sheet"""
        # Title
        worksheet['A1'].font = Font(bold=True, size=12)
        worksheet['B1'].font = Font(bold=True, size=12)
        
        # Highlight key metrics
        highlight_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row in range(2, worksheet.max_row + 1):
            worksheet[f'A{row}'].fill = highlight_fill
        
        # Auto-adjust widths
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 20
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get current statistics from the log"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Metrics')
            
            if df.empty:
                return {'message': 'No data available'}
            
            stats = {
                'total_sessions': len(df),
                'total_files': df['File_Name'].nunique(),
                'avg_processing_time': df['Processing_Time_Sec'].mean(),
                'avg_response_time': df['Response_Time_Sec'].mean(),
                'avg_total_runtime': df['Total_Runtime_Sec'].mean(),
                'success_rate': (df['Success'].sum() / len(df)) * 100,
                'total_words_processed': df['Words_Count'].sum(),
                'total_chunks_created': df['Chunks_Count'].sum(),
                'avg_words_per_file': df['Words_Count'].mean(),
                'avg_chunks_per_file': df['Chunks_Count'].mean(),
                'file_types': df['File_Type'].value_counts().to_dict(),
                'recent_sessions': df.tail(5).to_dict('records')
            }
            
            return stats
            
        except Exception as e:
            logger.error(f"Error getting statistics: {e}")
            return {'error': str(e)}
    
    def print_statistics(self):
        """Print formatted statistics to console"""
        stats = self.get_statistics()
        
        if 'error' in stats:
            print(f"Error: {stats['error']}")
            return
        
        if 'message' in stats:
            print(stats['message'])
            return
        
        print("\n" + "="*60)
        print("📊 CHAT METRICS STATISTICS")
        print("="*60)
        print(f"Total Sessions:           {stats['total_sessions']}")
        print(f"Total Files Processed:    {stats['total_files']}")
        print(f"Total Words Processed:    {stats['total_words_processed']:,}")
        print(f"Total Chunks Created:     {stats['total_chunks_created']:,}")
        print(f"\nAverages:")
        print(f"  Processing Time:        {stats['avg_processing_time']:.3f} sec")
        print(f"  Response Time:          {stats['avg_response_time']:.3f} sec")
        print(f"  Total Runtime:          {stats['avg_total_runtime']:.3f} sec")
        print(f"  Words per File:         {stats['avg_words_per_file']:.0f}")
        print(f"  Chunks per File:        {stats['avg_chunks_per_file']:.0f}")
        print(f"\nSuccess Rate:             {stats['success_rate']:.1f}%")
        print(f"\nFile Types:")
        for file_type, count in stats['file_types'].items():
            print(f"  {file_type}: {count}")
        print("="*60 + "\n")
    
    def export_summary_report(self, output_path: str = None):
        """Export a detailed summary report"""
        if output_path is None:
            output_path = f"metrics_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Metrics')
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Overall metrics
                df.to_excel(writer, sheet_name='All Data', index=False)
                
                # By file type
                file_type_stats = df.groupby('File_Type').agg({
                    'Session_ID': 'count',
                    'Words_Count': 'mean',
                    'Chunks_Count': 'mean',
                    'Processing_Time_Sec': 'mean',
                    'Response_Time_Sec': 'mean',
                    'Success': 'mean'
                }).reset_index()
                file_type_stats.columns = ['File Type', 'Count', 'Avg Words', 
                                          'Avg Chunks', 'Avg Processing', 
                                          'Avg Response', 'Success Rate']
                file_type_stats.to_excel(writer, sheet_name='By File Type', index=False)
                
                # Daily statistics
                df['Date'] = pd.to_datetime(df['Timestamp']).dt.date
                daily_stats = df.groupby('Date').agg({
                    'Session_ID': 'count',
                    'Words_Count': 'sum',
                    'Total_Runtime_Sec': 'sum'
                }).reset_index()
                daily_stats.columns = ['Date', 'Sessions', 'Total Words', 'Total Runtime']
                daily_stats.to_excel(writer, sheet_name='Daily Stats', index=False)
            
            logger.info(f"Summary report exported to {output_path}")
            print(f"✅ Summary report exported to {output_path}")
            
        except Exception as e:
            logger.error(f"Error exporting summary: {e}")
            print(f"❌ Error: {e}")


# =============================================================================
# USAGE EXAMPLES
# =============================================================================

def example_usage():
    """Example usage of the MetricsLogger"""
    
    # Initialize logger
    logger = ChatMetricsLogger("chat_metrics_log.xlsx")
    
    # Example 1: Track file upload and processing
    session_id = "session_123"
    logger.start_session(session_id)
    
    # Track file loading
    with logger.track_phase('file_load'):
        time.sleep(0.5)  # Simulate file loading
    
    # Log file info
    logger.log_file_info(
        file_name="example.pdf",
        file_type="pdf",
        file_size_bytes=1024 * 500,  # 500 KB
        words_count=5000,
        chunks_count=25
    )
    
    # Track processing
    with logger.track_phase('processing'):
        time.sleep(1.0)  # Simulate processing
    
    # Track embedding
    with logger.track_phase('embedding'):
        time.sleep(0.3)  # Simulate embedding generation
    
    # Track response generation
    with logger.track_phase('response_generation'):
        time.sleep(0.7)  # Simulate LLM response
    
    # Log query info
    logger.log_query_info(
        query_text="What are the main findings?",
        response_length=250,
        sources_retrieved=5,
        confidence_level="high"
    )
    
    # End session and save
    logger.end_session()
    
    # Print statistics
    logger.print_statistics()


if __name__ == "__main__":
    example_usage()